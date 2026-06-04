"""人员管理:列表、导入、删除、更新、批量导入、从设备移除"""

import io
import os
import shutil
from datetime import datetime, timedelta

from flask import Blueprint, request
from openpyxl import load_workbook

import db

from routes._app import BASE, manager
from routes.helpers import (
    ok,
    err,
    log_audit,
    get_current_user,
    load_devices,
    load_persons_for_device,
    load_persons_for_devices,
    _access_devices,
    _find_access_device_by_id,
    require_permission,
)


bp = Blueprint("persons", __name__)


@bp.route("/api/persons", methods=["GET"])
def get_persons():
    try:
        require_permission("page.persons")
    except Exception as e:
        return err(str(e), 403)
    device_id = request.args.get("device_id")
    page = request.args.get("page", type=int)
    per_page = request.args.get("per_page", 50, type=int)
    search = request.args.get("search", "").strip() or None
    search_mode = request.args.get("search_mode", "fuzzy")  # 'exact' 或 'fuzzy'
    username = get_current_user()
    if device_id:
        did = int(device_id)
        if not _find_access_device_by_id(did, username):
            return err("设备不存在或不是门禁设备", 404)
        if page:
            result = load_persons_for_devices([did], page=page, per_page=per_page, search=search, search_mode=search_mode)
        else:
            result = load_persons_for_device(did)
    else:
        # 全部设备：只返回当前用户设备中的人员（本地查找，按用户设备过滤）
        user_devices = _access_devices(load_devices(username))
        user_device_ids = {d["id"] for d in user_devices}
        if page:
            result = load_persons_for_devices(user_device_ids, page=page, per_page=per_page, search=search, search_mode=search_mode)
        else:
            result = load_persons_for_devices(user_device_ids)
    if page:
        persons, total = result
        return ok({"items": persons, "total": total, "page": page, "per_page": per_page})
    return ok(result)


@bp.route("/api/persons/import", methods=["POST"])
def import_persons():
    try:
        require_permission("page.persons")
    except Exception as e:
        return err(str(e), 403)
    data = request.get_json() or {}
    device_id = data.get("device_id")          # 全局ID
    new_persons = data.get("persons", [])
    no_purge = bool(data.get("no_purge", False))  # 非全量导入时传 True，跳过清理当前设备其他人员
    default_has_face = bool(data.get("default_has_face", not no_purge))  # 设备拉取/同步默认有人脸；手动新增(no_purge)默认不预设
    detect_orphans = bool(data.get("detect_orphans", False))
    detect_conflicts = bool(data.get("detect_conflicts", False))
    if device_id is None or not isinstance(new_persons, list):
        return err("参数错误")
    try:
        did = int(device_id)
    except (TypeError, ValueError):
        return err("设备ID错误")
    username = get_current_user()
    access_devices = _access_devices(load_devices(username))
    access_device_ids = {int(d["id"]) for d in access_devices}
    if did not in access_device_ids:
        return err("设备不存在或不是门禁设备", 404)
    for np in new_persons:
        if not isinstance(np, dict) or not np.get("user_id"):
            return err("人员数据错误")
        raw_doors = np.get("doors")
        if isinstance(raw_doors, list) and any(int(x) not in access_device_ids for x in raw_doors):
            return err("人员数据包含非门禁设备")

    result = db.import_persons_for_device(
        did, new_persons,
        no_purge=no_purge,
        default_has_face=default_has_face,
        detect_orphans=detect_orphans,
        detect_conflicts=detect_conflicts,
    )
    orphans = result.get("orphans", [])
    conflicts = result.get("conflicts", [])
    if detect_conflicts and conflicts:
        # 在冲突条目中补上设备名称
        dev_map = {d["id"]: d.get("name", d.get("ip", f"设备{d['id']}")) for d in access_devices}
        for c in conflicts:
            did_conf = c.get("from_device")
            if did_conf:
                c["device_name"] = dev_map.get(did_conf, f"设备{did_conf}")
    log_audit("import_persons", target=f"device_id={did}", detail={
        "count": len(new_persons),
        "no_purge": no_purge,
        "default_has_face": default_has_face,
        "orphans": len(orphans),
        "conflicts": len(conflicts),
    })
    payload = {"count": len(new_persons)}
    if detect_orphans:
        payload["orphans"] = orphans
    if detect_conflicts:
        payload["conflicts"] = conflicts
    return ok(payload)


@bp.route("/api/persons/<user_id>", methods=["PUT"])
def update_person(user_id):
    try:
        require_permission("page.persons")
    except Exception as e:
        return err(str(e), 403)
    data = request.get_json() or {}
    if not isinstance(data, dict):
        return err("参数错误")
    existing = db.load_person(user_id)
    if existing is None:
        return err("人员不存在", 404)
    username = get_current_user()
    managed_device_ids = [int(d["id"]) for d in _access_devices(load_devices(username))]
    try:
        updated = db.update_person_for_devices(user_id, data, managed_device_ids)
    except (TypeError, ValueError):
        return err("设备ID错误")
    if updated is None:
        return err("人员不存在", 404)
    log_audit("update_person", target=user_id, detail={"name": updated.get("name"), "doors": updated.get("doors")})
    return ok(updated)


@bp.route("/api/batch_import", methods=["POST"])
def batch_import():
    try:
        require_permission("page.persons")
        username = get_current_user()
    except Exception as e:
        log_audit("batch_import", result="fail", error=str(e))
        return err(str(e), 401)

    if 'files' not in request.files:
        log_audit("batch_import", result="fail", error="请选择文件夹上传")
        return err("请选择文件夹上传")

    uploaded_files = request.files.getlist('files')

    user_devices = _access_devices(load_devices(username))

    excel_data = None
    images = {}
    for f in uploaded_files:
        basename = os.path.basename(f.filename)
        if basename == 'user.xlsx':
            excel_data = f.read()
        elif f.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp')):
            images[basename] = f.read()

    if excel_data is None:
        log_audit("batch_import", result="fail", error="文件夹中未找到 user.xlsx 文件")
        return err("文件夹中未找到 user.xlsx 文件")

    try:
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            log_audit("batch_import", result="fail", error="Excel 数据不足（需要第一行说明、第二行列名、至少一行数据）")
            return err("Excel 数据不足（需要第一行说明、第二行列名、至少一行数据）")
        headers = [str(c).strip() if c else '' for c in rows[1]]
        required = ['用户编号', '姓名']
        for r in required:
            if r not in headers:
                log_audit("batch_import", result="fail", error=f"Excel 缺少必需列: {r}")
                return err(f"Excel 缺少必需列: {r}")
        data_rows = rows[2:]
    except Exception as e:
        log_audit("batch_import", result="fail", error=f"Excel 解析失败: {e}")
        return err(f"Excel 解析失败: {e}")

    total = 0
    success = 0
    fail = 0
    face_success = 0
    face_fail = 0
    details = []          # 存储每条记录的结果

    for row in data_rows:
        # 跳过完全为空的行（所有单元格均为 None 或空字符串）
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        # 提取用户编号和姓名，如果无效则跳过该行
        try:
            user_id = str(row[headers.index('用户编号')]).strip()
            name = str(row[headers.index('姓名')]).strip()
        except (ValueError, IndexError):
            continue
        if not user_id or not name:
            continue

        total += 1

        valid_begin = "2000-01-01"
        valid_end = None
        if '有效期结束' in headers:
            ve = row[headers.index('有效期结束')]
            if ve is not None:
                ve_str = str(ve).strip()
                if ve_str and ve_str != 'None':
                    # 尝试多种日期格式，转换为标准 YYYY-MM-DD
                    for fmt in ("%Y-%m-%d", "%Y%m%d", "%Y/%m/%d", "%Y.%m.%d"):
                        try:
                            valid_end = datetime.strptime(ve_str, fmt).strftime("%Y-%m-%d")
                            break
                        except ValueError:
                            continue
        # 如果未能成功解析或未填写，默认 10 年
        if not valid_end:
            valid_end = (datetime.now() + timedelta(days=3650)).strftime("%Y-%m-%d")

        face_filename = ''
        if '人脸图片名称' in headers:
            idx = headers.index('人脸图片名称')
            raw = row[idx]
            face_filename = str(raw).strip() if raw else ''

        door_names = []
        if '门' in headers:
            idx = headers.index('门')
            raw = row[idx]
            if raw:
                door_names = [d.strip() for d in str(raw).split(',') if d.strip()]

        door_ids = []
        resolved_doors = []     # 用于详情显示的门名称列表
        if door_names:
            # 表格中指定了设备名称,逐个精确匹配,避免子串模糊命中错误设备。
            # 匹配优先级:精确名称 > 精确 ip:port > 精确 ip。
            # 多匹配视为歧义直接报错,绝不下发到任意一台。
            name_index = {d['name']: d for d in user_devices}
            endpoint_index = {f"{d['ip']}:{d['port']}": d for d in user_devices}
            ip_index = {}
            for d in user_devices:
                ip_index.setdefault(d['ip'], []).append(d)

            for dn in door_names:
                dev = None
                if dn in name_index:
                    dev = name_index[dn]
                elif dn in endpoint_index:
                    dev = endpoint_index[dn]
                elif dn in ip_index:
                    if len(ip_index[dn]) == 1:
                        dev = ip_index[dn][0]
                    else:
                        resolved_doors.append(f"{dn}(歧义:同 IP 多设备)")
                        continue
                if dev:
                    if dev['id'] not in door_ids:  # 避免重复
                        door_ids.append(dev['id'])
                        resolved_doors.append(dev['name'])
                else:
                    resolved_doors.append(f"{dn}(未找到)")
        else:
            # 未指定任何设备，不执行添加，直接计入失败
            fail += 1
            details.append({
                "user_id": user_id,
                "name": name,
                "doors": "未指定设备",
                "status": "失败",
                "face_result": "—",
                "error": "Excel中未填写目标设备"
            })
            continue    # 跳过该行，不再使用当前选中设备

        if not door_ids:
            # 所有指定设备都未找到
            fail += 1
            details.append({
                "user_id": user_id,
                "name": name,
                "doors": ", ".join(resolved_doors) if resolved_doors else "无",
                "status": "失败",
                "face_result": "—",
                "error": "未找到有效设备"
            })
            continue

        # 每行的人员添加可能成功部分门，我们记录整体状态
        row_success = 0
        row_fail = 0
        row_face_ok = 0
        row_face_fail = 0
        error_messages = []

        for did in door_ids:
            dev = next((d for d in user_devices if d['id'] == did), None)
            if not dev:
                row_fail += 1
                error_messages.append(f"设备{did}不存在")
                continue

            try:
                manager.get(dev).add_user(user_id, name, valid_begin=valid_begin, valid_end=valid_end, status=0, doors=[0])

                existing_person = db.load_person(user_id)
                has_face_for_device = bool((existing_person or {}).get("has_face", {}).get(str(did), False))
                person_payload = existing_person or {
                    "user_id": user_id,
                    "name": name,
                    "valid_begin": valid_begin,
                    "valid_end": valid_end,
                    "doors": [],
                    "status": {},
                    "has_face": {},
                }
                person_payload["name"] = name
                person_payload["valid_begin"] = valid_begin
                person_payload["valid_end"] = valid_end
                if did not in person_payload.setdefault("doors", []):
                    person_payload["doors"].append(did)
                person_payload.setdefault("status", {})[str(did)] = 0
                person_payload.setdefault("has_face", {}).setdefault(str(did), has_face_for_device)
                db.upsert_person(person_payload)
                row_success += 1

                # 人脸下发
                if face_filename and face_filename in images:
                    # 检查该设备是否已有脸，已有则跳过下发
                    if has_face_for_device:
                        row_face_ok += 1  # 已有，视为成功
                    else:
                        img_bytes = images[face_filename]
                        tmp_path = os.path.join(BASE, f"_batch_{user_id}_{did}.jpg")
                        with open(tmp_path, 'wb') as f_img:
                            f_img.write(img_bytes)
                        try:
                            manager.get(dev).add_face(user_id, tmp_path)
                            # 缓存到本地
                            face_dir = os.path.join(BASE, "faces")
                            os.makedirs(face_dir, exist_ok=True)
                            shutil.copy(tmp_path, os.path.join(face_dir, f"{user_id}.jpg"))
                            db.set_person_face(user_id, did, True)
                            row_face_ok += 1
                        except Exception as e:
                            row_face_fail += 1
                            dev_label = dev.get('name') or f"{dev.get('ip')}:{dev.get('port')}" or str(did)
                            error_messages.append(f"设备{dev_label}人脸下发失败: {e}")
                        finally:
                            if os.path.exists(tmp_path):
                                os.remove(tmp_path)
            except Exception as e:
                row_fail += 1
                error_messages.append(str(e))

        # 汇总该行结果
        status_text = "成功"
        if row_success == 0:
            status_text = "失败"
        elif row_fail > 0:
            status_text = "部分成功"

        face_text = "—"
        if face_filename:
            if row_face_ok == len(door_ids):
                face_text = "全部成功"
            elif row_face_ok > 0:
                face_text = f"部分成功({row_face_ok}/{len(door_ids)})"
            else:
                face_text = "失败"

        if row_success > 0:
            success += row_success
        if row_fail > 0:
            fail += row_fail
        face_success += row_face_ok
        face_fail += row_face_fail

        details.append({
            "user_id": user_id,
            "name": name,
            "doors": ", ".join(resolved_doors) if resolved_doors else "无",
            "status": status_text,
            "face_result": face_text,
            "error": "; ".join(error_messages) if error_messages else ""
        })

    msg = f"处理 {total} 条，成功添加 {success} 人"
    if face_success + face_fail > 0:
        msg += f"，人脸下发 {face_success} 成功 / {face_fail} 失败"
    log_audit("batch_import", detail={"total": total, "success": success, "fail": fail, "face_success": face_success, "face_fail": face_fail})
    return ok({
        "total": total,
        "success": success,
        "fail": fail,
        "face_success": face_success,
        "face_fail": face_fail,
        "msg": msg,
        "details": details
    })


# 从指定设备移除人员关联（全局ID）
@bp.route("/api/persons/<user_id>/devices/<int:device_id>", methods=["DELETE"])
def remove_person_from_device(user_id, device_id):
    try:
        require_permission("page.persons")
    except Exception as e:
        return err(str(e), 403)
    username = get_current_user()
    if not _find_access_device_by_id(device_id, username):
        return err("设备不存在或不是门禁设备", 404)
    result = db.remove_person_device(user_id, device_id)
    if result and result.get("status") == "ok":
        log_audit("remove_person_device", target=f"{user_id}->device_{device_id}")
        return ok({"removed_device": device_id, "remaining_doors": result["remaining_doors"]})
    if result and result.get("status") == "not_on_device":
        return err("该人员未关联此设备", 404)
    return err("人员不存在", 404)
